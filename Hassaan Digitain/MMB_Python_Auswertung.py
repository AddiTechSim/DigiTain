import numpy as np
import matplotlib.pyplot as plt
import csv
from scipy.integrate import trapz
import glob
import openpyxl

######################################################################################################################
######################################### Kolletition of needed Functions ############################################
######################################################################################################################

def read_csv(file_path):

    #Read data from CSV file, convert "," to "." and extract displacement and force values.

    displacement = []
    force = []
    try:
        with open(file_path, 'r') as file:
            reader = csv.reader(file, delimiter=';')
            next(reader)  # Skip header
            for row in reader:
                    # Replace commas with periods
                    row = [entry.replace(',', '.') for entry in row]
                    if float(row[1]) >= 0:
                        displacement.append(float(row[2]))  # Column 3 for Displacement
                        force.append(float(row[1]))  # Column 2 for Force
    except FileNotFoundError:
        print("Error: File not found.")
    except Exception as e:
            if not force and not displacement:
                with open(file_path, 'r') as csvfile:
                    csvreader = csv.reader(csvfile)
                    header = next(csvreader)  # Skip the header row
                    for row in csvreader:
                        displacement.append(float(row[0]))
                        force.append(float(row[1]))
    return displacement, force


def analyze_data(displacement, force):
    # Analyze the data for the Maximum Measured Load and calculate the linear equation between 20% and 70% load.

    max_force_index = np.argmax(force)
    max_force = [displacement[max_force_index], force[max_force_index]]
    start_index = int(0.2 * max_force_index)
    end_index = int(0.7 * max_force_index)
    force_range = force[start_index:end_index]
    displacement_range = displacement[start_index:end_index]
    slope = np.diff(force_range) / np.diff(displacement_range)
    average_slope = np.mean(slope)
    c = force_range[0] - average_slope * displacement_range[0]
    return max_force, [average_slope, c]


# Find intersections between the reduced slope line and the force-displacement curve
def find_intersections(points, p0, p1):
    # Calculate gradient and intercept of the line
    b = (p1[1] - p0[1]) / (p1[0] - p0[0])
    a = p0[1] - b * p0[0]
    
    # Calculate distance of y value from the line
    B = (a + points[:, 0] * b) - points[:, 1]
    
    # Find indices of points where the next point is on the other side of the line
    ix = np.where(B[1:] * B[:-1] < 0)[0]
    
    # Calculate crossing points
    cross_points = np.zeros((len(ix), 2))
    cross_points[:, 0] = points[ix, 0] + (B[ix] / (B[ix] - B[ix + 1])) * (points[ix + 1, 0] - points[ix, 0])
    cross_points[:, 1] = points[ix, 1] + (B[ix] / (B[ix] - B[ix + 1])) * (points[ix + 1, 1] - points[ix, 1])
    
    # Find the x-value at the maximum y-value
    max_y_index = np.argmax(cross_points[:, 1])
    x_at_max_y = cross_points[max_y_index, 0]
    max_y = cross_points[max_y_index, 1]
    max_y_value =[x_at_max_y,max_y]

    return cross_points, max_y_value

def calculate_fracture_toughness(displacement, force):
    #currently a dummy funktion ment for further development
    #Calculate fracture toughness using the area under the force-displacement curve.

    area_under_curve = trapz(force, displacement)
    fracture_toughness = (area_under_curve / (2 * np.pi)) ** 0.5
    return fracture_toughness

def get_value_of_X(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Extracting values from the Excel sheet (assuming X is in cell A1)
        h = sheet['B2'].value
        E22 = sheet['B3'].value
        G13 = sheet['B4'].value
        r = 1.18*((h*E22)**0.5)/G13
        X_value = (h/(11*G13)*(3-2*(r/(1+r))**2))**0.5

        return X_value

    except Exception as e:
        print(f"An error occurred: {e}")
        return None
def calc_E1f(file_path,specimen_Name,X,m,P):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Default values in case the specimen name is not found
        b = h = a0 = L = c = Csys = None

        # Extracting values from the Excel sheet (assuming X is in cell A1)

        # Search for the string "csv_name" within the range A13:A100
        for row in sheet.iter_rows(min_row=13, max_row=100, min_col=1, max_col=1, values_only=False):
            if row[0].value == specimen_Name:
                row_number = row[0].row
                # If the string is found, read values from columns B and C in the same row
                b = sheet.cell(row=row_number, column=2).value
                h = (sheet.cell(row=row_number, column=3).value)/2
                a0 =sheet.cell(row=row_number, column=4).value
                L = sheet.cell(row=row_number, column=5).value
                c = sheet.cell(row=row_number, column=6).value
                Csys = sheet.cell(row=row_number, column=7).value

            if b is None or h is None or a0 is None or L is None or c is None or Csys is None:
                b = sheet.cell(row=13, column=2).value
                h = (sheet.cell(row=13, column=3).value) / 2
                a0 = sheet.cell(row=13, column=4).value
                L = sheet.cell(row=13, column=5).value
                c = sheet.cell(row=13, column=6).value
                Csys = sheet.cell(row=13, column=7).value


        E1f =(8*(a0+X*h)**3*(3*c-L)**2+(6*(a0+0.42*h*X)**3+4*L**3)*(c+L)**2)/(16*L**2*b*h**3*(1/m-Csys))
        #=(8*(a0+X*h)^3*((3*c-L)^2)+(6*(a0+0.43*X*h)^(3)+4*(L^3))*(c+L)^2)/(16*(L^2)*b*(h^3)*((1/m)-Csys))


        GI =(12*P**2*(3*c-L)**2)/(16*b**2*h**3*L**2*E1f)*(a0+X*h)**2


        GII =(9*P**2*(c+L)**2)/(16*b**2*h**3*L**2*E1f)*(a0+0.42*X*h)**2

        Gc = GI + GII

        MMR = GII/Gc

        return Gc,MMR

    except Exception as e:
        print(f"An error occurred: {e}")
        return None
######################################################################################################################
######################################### Start of the actual Main Skript ############################################
######################################################################################################################

# Get a list of all CSV files in the current directory
csv_files = glob.glob("*.csv")

# Remove 'output.csv' from the list of CSV files
csv_files = [file for file in csv_files if file != 'output.csv']

xlsx_file = "Probendaten.xlsx"

X = get_value_of_X(xlsx_file)

GC_List = [['Proben_name','MMR','Gc_NL','Gc_5%/Max','Force_NL','Force_5%/Max']]

# Iterate over each CSV file
for csv_file_path in csv_files:

    # Read data from CSV
    displacement, force = read_csv(csv_file_path)

    # Analyze the data
    max_force, linearization = analyze_data(displacement, force)

    # Calculate a line with a 5% reduced slope
    reduced_slope = 0.95 * linearization[0]
    x_line = np.array([0, max(displacement)])
    y_line = reduced_slope * x_line+linearization[1]

    # create line for plotting the linearization
    x_org = np.array([0, max(displacement)])
    y_org = linearization[0] * x_org+linearization[1]

    points = np.column_stack((displacement, force))
    max_displacement = max(displacement)

    #find the intersections and the Max Force intersection of the 5% reduced slope and the Measurement
    intersections,max_intersection_force = find_intersections(points,[0,linearization[1]],[max_displacement,max_displacement*reduced_slope+linearization[1]])

    #find the Max Force intersection of the linearization and the Measurement
    intersections2,NoneLinear = find_intersections(points,[0,linearization[1]],[max_displacement,max_displacement*linearization[0]+linearization[1]])

    if max_intersection_force[0] < max_force[0]:
        Force_5Max = max_intersection_force[1]
    else:
        Force_5Max = max_force[1]

    print("5%/Max Force:", Force_5Max)
    print("NL Force:", NoneLinear[1])

    # Calculate fracture toughness
    Gc_5Max = calc_E1f(xlsx_file,csv_file_path[:-4],X,linearization[0],Force_5Max)[0]


    Gc_NL = calc_E1f(xlsx_file,csv_file_path[:-4],X,linearization[0],NoneLinear[1])


    GC_List.append([csv_file_path[:-4],Gc_NL[1],Gc_NL[0],Gc_5Max,NoneLinear[1],Force_5Max])

    #fracture_toughness = calculate_fracture_toughness(displacement, force)


    # Plot force-displacement curve, the reduced slope line, and the intersection points
    plt.plot(displacement, force, color='black', label='Force-Displacement Curve')
    plt.plot(x_org, y_org, linestyle=':', color='red', label='linearization')
    plt.plot(x_line, y_line, linestyle='--', color='red', label='5% reduced Slope')
    plt.scatter(max_intersection_force[0],max_intersection_force[1],marker='x', color='green', label=f'5% = {str(round(max_intersection_force[1], 2))}N')
    plt.scatter(max_force[0],max_force[1],marker='x', color='blue', label=f'Max = {str(round(max_force[1], 2))}N')
    plt.scatter(NoneLinear[0],NoneLinear[1],marker='x', color='darkorange', label=f'NL = {str(round(NoneLinear[1], 2))}N')
    plt.xlabel('Displacement')
    plt.ylabel('Force')
    plt.title(f'Force-Displacement Curve of {csv_file_path[:-4]}')
    plt.legend()
    plt.grid(True)
    plt.savefig(f'{csv_file_path[:-4]}.svg')
    plt.savefig(f'{csv_file_path[:-4]}.png', format="png", dpi=400)
    plt.cla()
print(GC_List)

with open('output.csv', 'w', newline='') as f:
    writer = csv.writer(f, delimiter=';')
    writer.writerows(GC_List)

